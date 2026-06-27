"""
DataForge Main Application GUI

Modern tkinter GUI with ttkbootstrap theming.
Provides: file upload, cleaning review, DDL/DML output, Power BI export,
          and live Chart.js dashboard via embedded Flask server.
"""

import sys
import os
import threading
import webbrowser
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List

try:
    import ttkbootstrap as ttk
    from ttkbootstrap.constants import *
    from ttkbootstrap.dialogs import Messagebox
    from ttkbootstrap.scrolled import ScrolledText
    HAS_TTKBOOTSTRAP = True
except ImportError:
    import tkinter as tk
    import tkinter.ttk as ttk
    from tkinter import messagebox
    from tkinter.scrolled import ScrolledText
    HAS_TTKBOOTSTRAP = False

    class _MB:
        @staticmethod
        def show_error(msg, title=""): messagebox.showerror(title, msg)
        @staticmethod
        def show_warning(msg, title=""): messagebox.showwarning(title, msg)
        @staticmethod
        def show_info(msg, title=""): messagebox.showinfo(title, msg)
    Messagebox = _MB()

    BOTH = "both"
    LEFT = "left"
    RIGHT = "right"
    W = "w"
    X = "x"
    Y = "y"
    VERTICAL = "vertical"
    HORIZONTAL = "horizontal"

from tkinter import filedialog, StringVar, BooleanVar, IntVar

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core.schema_engine import SchemaEngine
from core.cleaner import DataCleaner
from core.project_manager import ProjectManager
from utils.file_handler import FileHandler
from utils.formatted_exporter import FormattedExcelExporter
from bi.powerbi_exporter import PowerBIExporter
from db.duckdb_manager import DuckDBManager
from dashboard.server import DashboardServer
from dashboard.chart_suggester import ChartSuggester


class DataForgeApp:
    """Main DataForge application window."""

    APP_NAME = "DataForge"
    APP_VERSION = "2.0.0"
    DIALECTS = ['snowflake', 'sqlserver', 'mysql', 'postgresql', 'oracle']
    DML_LIMITS = [100, 500, 1000, 999999]

    def __init__(self):
        if HAS_TTKBOOTSTRAP:
            self.root = ttk.Window(themename="darkly")
        else:
            import tkinter as tk
            self.root = tk.Tk()

        self.root.title(f"{self.APP_NAME} v{self.APP_VERSION} — Power BI Edition")
        self.root.geometry("1400x900")
        self.root.minsize(1200, 700)

        self.file_handler = FileHandler()
        self.schema_engine: Optional[SchemaEngine] = None
        self.cleaner = DataCleaner()
        self.project_manager = ProjectManager()
        self.formatted_exporter = FormattedExcelExporter()
        self.duckdb: Optional[DuckDBManager] = None

        self.current_file: Optional[str] = None
        self.headers: List[str] = []
        self.raw_rows: List[List[Any]] = []
        self.cleaned_rows: List[List[Any]] = []
        self.file_info: Dict[str, Any] = {}
        self.all_sheets: Dict[str, Any] = {}

        self.table_name_var = StringVar(value="my_table")
        self.dialect_var = StringVar(value="snowflake")
        self.nullable_var = BooleanVar(value=True)
        self.dml_limit_var = IntVar(value=1000)
        self.sheet_var = StringVar()
        self.multi_sheet_var = BooleanVar(value=False)

        self._build_ui()
        self._load_preferences()

    def _build_ui(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_rowconfigure(1, weight=1)

        # === TOOLBAR ===
        toolbar = ttk.Frame(main_frame)
        toolbar.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 10))

        ttk.Label(toolbar, text="DataForge", font=("Helvetica", 18, "bold")).pack(side=LEFT)
        ttk.Label(toolbar, text=f"v{self.APP_VERSION}", font=("Helvetica", 10)).pack(side=LEFT, padx=(5, 20))

        ttk.Button(toolbar, text="Browse File", command=self._browse_file).pack(side=LEFT, padx=5)
        ttk.Button(toolbar, text="Load Demo Data", command=self._load_demo).pack(side=LEFT, padx=5)
        ttk.Button(toolbar, text="Recent Projects", command=self._show_recent).pack(side=LEFT, padx=5)

        if HAS_TTKBOOTSTRAP:
            ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=10)
        ttk.Button(toolbar, text="Save Project", command=self._save_project).pack(side=LEFT, padx=5)

        # === LEFT PANEL: Settings ===
        left_panel = ttk.LabelFrame(main_frame, text="Settings")
        left_panel.grid(row=1, column=0, sticky="ns", padx=(0, 10), pady=5)
        left_panel.configure(width=260)
        left_panel.grid_propagate(False)

        left_inner = ttk.Frame(left_panel)
        left_inner.pack(fill=BOTH, expand=True, padx=10, pady=10)

        ttk.Label(left_inner, text="Table Name:").pack(anchor=W, pady=(0, 2))
        ttk.Entry(left_inner, textvariable=self.table_name_var, width=28).pack(fill=X, pady=(0, 10))

        ttk.Label(left_inner, text="SQL Dialect:").pack(anchor=W, pady=(0, 2))
        dialect_combo = ttk.Combobox(left_inner, textvariable=self.dialect_var,
                                     values=self.DIALECTS, state="readonly", width=26)
        dialect_combo.pack(fill=X, pady=(0, 10))
        dialect_combo.set("snowflake")

        ttk.Label(left_inner, text="Sheet:").pack(anchor=W, pady=(0, 2))
        self.sheet_combo = ttk.Combobox(left_inner, textvariable=self.sheet_var,
                                        state="readonly", width=26)
        self.sheet_combo.pack(fill=X, pady=(0, 10))
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_change)

        ttk.Checkbutton(left_inner, text="Process ALL sheets",
                        variable=self.multi_sheet_var).pack(anchor=W, pady=(0, 5))
        ttk.Checkbutton(left_inner, text="Columns NULLable by default",
                        variable=self.nullable_var).pack(anchor=W, pady=(0, 10))

        ttk.Label(left_inner, text="DML Row Limit:").pack(anchor=W, pady=(0, 2))
        limit_frame = ttk.Frame(left_inner)
        limit_frame.pack(fill=X, pady=(0, 10))
        for val in self.DML_LIMITS:
            label = "All" if val == 999999 else str(val)
            ttk.Radiobutton(limit_frame, text=label, variable=self.dml_limit_var,
                            value=val).pack(side=LEFT, padx=3)

        ttk.Button(left_inner, text="Generate DDL + DML",
                   command=self._generate).pack(fill=X, pady=(10, 5))

        self.file_info_label = ttk.Label(left_inner, text="No file loaded",
                                          wraplength=230, justify=LEFT)
        self.file_info_label.pack(anchor=W, pady=(20, 0))

        # === CENTER: Notebook Tabs ===
        center_panel = ttk.Frame(main_frame)
        center_panel.grid(row=1, column=1, sticky="nsew", pady=5)

        self.notebook = ttk.Notebook(center_panel)
        self.notebook.pack(fill=BOTH, expand=True)

        tab_configs = [
            ("ddl_text", "DDL Output"),
            ("dml_text", "DML Output"),
            ("dict_text", "Data Dictionary"),
            ("map_text", "Column Name Map"),
            ("audit_text", "Cleaning Audit"),
            ("preview_text", "Row Preview"),
            ("sheets_text", "Sheet Preview"),
        ]

        for attr, label in tab_configs:
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text=label)
            widget = ScrolledText(tab, wrap="word", height=30, font=("Courier New", 10), width=70)
            widget.pack(fill=BOTH, expand=True, padx=5, pady=5)
            setattr(self, attr, widget)

        # === RIGHT PANEL: Export ===
        right_panel = ttk.LabelFrame(main_frame, text="Export")
        right_panel.grid(row=1, column=2, sticky="ns", padx=(10, 0), pady=5)
        right_panel.configure(width=200)
        right_panel.grid_propagate(False)

        right_inner = ttk.Frame(right_panel)
        right_inner.pack(fill=BOTH, expand=True, padx=8, pady=8)

        for label, cmd in [
            ("Copy DDL", self._copy_ddl),
            ("Copy DML", self._copy_dml),
            ("Save .sql (DDL)", self._save_ddl),
            ("Save DML .sql", self._save_dml),
            ("Save Staged CSV", self._save_csv),
            ("Save Dict .xlsx", self._save_dict_xlsx),
            ("Save Dict .csv", self._save_dict_csv),
            ("Export Styled .xlsx", self._export_styled_excel),
        ]:
            ttk.Button(right_inner, text=label, command=cmd).pack(fill=X, pady=2)

        if HAS_TTKBOOTSTRAP:
            ttk.Separator(right_inner, orient=HORIZONTAL).pack(fill=X, pady=8)
        else:
            ttk.Separator(right_inner).pack(fill=X, pady=8)

        if HAS_TTKBOOTSTRAP:
            ttk.Button(right_inner, text="Power BI Export",
                       command=self._powerbi_export,
                       bootstyle="primary").pack(fill=X, pady=2)
        else:
            ttk.Button(right_inner, text="Power BI Export",
                       command=self._powerbi_export).pack(fill=X, pady=2)

        ttk.Button(right_inner, text="Preview M Code",
                   command=self._preview_m_code).pack(fill=X, pady=2)
        ttk.Button(right_inner, text="Launch Dashboard",
                   command=self._launch_dashboard).pack(fill=X, pady=2)

        # === STATUS BAR ===
        self.status_var = StringVar(value="Ready — Load a file or click 'Load Demo Data'")
        status_bar = ttk.Frame(main_frame)
        status_bar.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(10, 0))
        if HAS_TTKBOOTSTRAP:
            ttk.Separator(status_bar, orient=HORIZONTAL).pack(fill=X)
        ttk.Label(status_bar, textvariable=self.status_var,
                  font=("Helvetica", 9)).pack(anchor=W, pady=5)

    # ── File Loading ──────────────────────────────────────────────────────────

    def _browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel or CSV file",
            filetypes=[("Excel files", "*.xlsx *.xls"),
                       ("CSV files", "*.csv"),
                       ("All files", "*.*")]
        )
        if file_path:
            self._load_file(file_path)

    def _load_file(self, file_path: str):
        try:
            self.status_var.set(f"Loading {Path(file_path).name}...")
            self.root.update_idletasks()

            self.file_info = self.file_handler.read_file(file_path)
            self.current_file = file_path

            if self.file_info['file_type'] == 'excel':
                self.sheet_combo['values'] = self.file_info['sheets']
                self.sheet_combo.set(self.file_info['default_sheet'])
                self.sheet_combo.configure(state="readonly")
                if len(self.file_info['sheets']) > 1:
                    self.all_sheets = self.file_handler.get_all_sheets(file_path)
                    self._display_sheet_preview()
            else:
                self.sheet_combo['values'] = ['Sheet1']
                self.sheet_combo.set('Sheet1')
                self.sheet_combo.configure(state="disabled")
                self.all_sheets = {}

            base_name = Path(file_path).stem
            self.table_name_var.set(self._clean_table_name(base_name))

            info_text = (f"File: {Path(file_path).name}\n"
                         f"Type: {self.file_info['file_type'].upper()}\n"
                         f"Size: {self.file_info.get('file_size_mb', 'N/A')} MB\n"
                         f"Columns: {self.file_info['total_columns']}\n")
            if self.file_info.get('total_rows'):
                info_text += f"Rows: {self.file_info['total_rows']}\n"
            if self.file_info.get('sheets'):
                info_text += f"Sheets: {len(self.file_info['sheets'])}"

            self.file_info_label.config(text=info_text)
            self.status_var.set(f"Loaded {Path(file_path).name} — click 'Generate DDL + DML'")

        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            Messagebox.show_error(f"Failed to load file:\n{str(e)}", "Error")

    def _display_sheet_preview(self):
        self.sheets_text.delete("1.0", "end")
        if not self.all_sheets:
            self.sheets_text.insert("end", "No multi-sheet data loaded.\n")
            return
        for sheet_name, (headers, rows) in self.all_sheets.items():
            self.sheets_text.insert("end", f"\n{'=' * 60}\n")
            self.sheets_text.insert("end", f"SHEET: {sheet_name}\n")
            self.sheets_text.insert("end", f"{'=' * 60}\n")
            self.sheets_text.insert("end", f"Columns: {len(headers)} | Rows: {len(rows)}\n")
            self.sheets_text.insert("end", f"Headers: {', '.join(headers)}\n\n")

    def _on_sheet_change(self, event=None):
        if self.current_file and self.file_info.get('file_type') == 'excel':
            sheet = self.sheet_var.get()
            if sheet in self.all_sheets:
                _, rows = self.all_sheets[sheet]
                self.status_var.set(f"Selected sheet: {sheet} ({len(rows)} rows)")

    def _load_demo(self):
        self.headers = ['Customer ID', 'Name', 'Sign-Up Date', 'Monthly Charges', 'Is Active', 'Region']
        self.raw_rows = [
            ['1', '  Acme Inc.  ', '03/15/2025', '120.50', 'yes', 'North'],
            ['2', 'Globex Corp.', '2025-04-20', '85.00', 'true', 'South'],
            ['3', 'NA', '05/01/2025', '0', 'no', 'East'],
            ['4', 'Initech LLC', '2025-06-10', '150.00', '1', 'West'],
            ['5', 'Umbrella Corp', '2025-01-15', '200.00', 'yes', 'North'],
        ]
        self.current_file = "DEMO"
        self.file_info = {'file_type': 'demo', 'total_columns': len(self.headers), 'total_rows': len(self.raw_rows)}
        self.table_name_var.set("demo_customers")
        self.file_info_label.config(text="DEMO DATA\n5 rows, 6 columns\nClick 'Generate DDL + DML'")
        self.status_var.set("Demo data loaded — click 'Generate DDL + DML'")

    # ── Generation ────────────────────────────────────────────────────────────

    def _generate(self):
        if not self.current_file:
            Messagebox.show_warning("Please load a file first.", "No File")
            return

        try:
            self.status_var.set("Generating...")
            self.root.update_idletasks()

            if self.multi_sheet_var.get() and self.all_sheets:
                self._generate_multi_sheet()
                return

            if self.current_file != "DEMO":
                sheet = self.sheet_var.get() if self.file_info['file_type'] == 'excel' else None
                self.headers, self.raw_rows = self.file_handler.load_sheet(self.current_file, sheet)

            self.cleaned_rows = self.cleaner.clean(self.headers, self.raw_rows)
            self.schema_engine = SchemaEngine(dialect=self.dialect_var.get())
            self.schema_engine.analyze(self.headers, self.cleaned_rows)
            self._display_outputs()

            # Load into DuckDB
            if self.duckdb:
                self.duckdb.close()
            self.duckdb = DuckDBManager()
            self.duckdb.load_data(
                [c['clean_name'] for c in self.schema_engine.columns],
                self.cleaned_rows,
                self.table_name_var.get()
            )

        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            Messagebox.show_error(f"Generation failed:\n{str(e)}", "Error")

    def _generate_multi_sheet(self):
        all_headers = []
        all_rows = []
        for sheet_name, (headers, rows) in self.all_sheets.items():
            if not all_headers:
                all_headers = headers
            all_rows.extend(rows)
        self.headers = all_headers
        self.raw_rows = all_rows
        self.cleaned_rows = self.cleaner.clean(self.headers, self.raw_rows)
        self.schema_engine = SchemaEngine(dialect=self.dialect_var.get())
        self.schema_engine.analyze(self.headers, self.cleaned_rows)
        self._display_outputs()

    def _display_outputs(self):
        table_name = self.table_name_var.get()
        nullable = self.nullable_var.get()

        ddl = self.schema_engine.generate_ddl(table_name, nullable)
        dml = self._generate_dml(table_name)
        dictionary = self.schema_engine.generate_data_dictionary()
        col_map = self.schema_engine.get_column_name_map()
        audit = self.cleaner.get_audit_log()
        preview = self.cleaner.get_preview(50)
        quality = self.schema_engine.get_quality_summary()

        self.ddl_text.delete("1.0", "end")
        self.ddl_text.insert("1.0", ddl)

        self.dml_text.delete("1.0", "end")
        self.dml_text.insert("1.0", dml)

        self.dict_text.delete("1.0", "end")
        self._format_dictionary(dictionary)

        self.map_text.delete("1.0", "end")
        for orig, clean in col_map.items():
            self.map_text.insert("end", f"{orig:30} -> {clean}\n")

        self.audit_text.delete("1.0", "end")
        self._format_audit(audit)

        self.preview_text.delete("1.0", "end")
        self._format_preview(preview)

        self.status_var.set(
            f"Generated! {quality['total_rows']} rows, {quality['total_columns']} cols | "
            f"Nulls: {quality['null_percentage']}% | "
            f"PK: {', '.join(quality['pk_candidates']) or 'None detected'}"
        )

    def _generate_dml(self, table_name: str) -> str:
        lines = [f"-- DML for {table_name}\n",
                 f"-- Generated by DataForge {self.APP_VERSION}\n",
                 f"-- Dialect: {self.dialect_var.get()}\n",
                 "=" * 60 + "\n\n"]

        limit = self.dml_limit_var.get()
        lines.append(f"-- INSERT statements (first {limit if limit != 999999 else 'all'} rows)\n")

        clean_names = [c['clean_name'] for c in self.schema_engine.columns]
        col_list = ", ".join(clean_names)

        for row in self.cleaned_rows[:limit]:
            values = []
            for val in row:
                if val is None:
                    values.append("NULL")
                elif isinstance(val, (int, float)):
                    values.append(str(val))
                else:
                    escaped = str(val).replace("'", "''")
                    values.append(f"'{escaped}'")
            lines.append(f"INSERT INTO {table_name} ({col_list}) VALUES ({', '.join(values)});\n")

        lines.append(f"\n-- Total: {min(limit, len(self.cleaned_rows))} rows\n")
        return "".join(lines)

    def _format_dictionary(self, dictionary: List[Dict]):
        self.dict_text.insert("end", f"{'Column':25} {'Type':25} {'Null%':>6} {'PK':>3} {'FK':>4} {'Sample'}\n")
        self.dict_text.insert("end", "=" * 100 + "\n")
        for col in dictionary:
            pk = "YES" if col['is_primary_key'] else ""
            fk = "HINT" if col['is_foreign_key_hint'] else ""
            self.dict_text.insert(
                "end",
                f"{col['column_name']:25} {col['data_type']:25} {col['null_percentage']:>6.1f} {pk:>3} {fk:>4} {col['sample_values']}\n"
            )

    def _format_audit(self, audit: List[Dict]):
        if not audit:
            self.audit_text.insert("end", "No cleaning actions performed.\n")
            return
        self.audit_text.insert("end", f"{'Action':25} {'Rows Affected':>12}  Description\n")
        self.audit_text.insert("end", "=" * 80 + "\n")
        for entry in audit:
            self.audit_text.insert(
                "end",
                f"{entry['action']:25} {entry.get('rows_affected', 0):>12}  {entry['description']}\n"
            )

    def _format_preview(self, preview: List[List[Any]]):
        if not self.schema_engine:
            return
        headers = [c['clean_name'] for c in self.schema_engine.columns]
        self.preview_text.insert("end", " | ".join(headers) + "\n")
        self.preview_text.insert("end", "=" * 100 + "\n")
        for row in preview:
            formatted = [str(v) if v is not None else "NULL" for v in row]
            self.preview_text.insert("end", " | ".join(formatted) + "\n")

    def _clean_table_name(self, name: str) -> str:
        import re
        clean = re.sub(r'[^a-zA-Z0-9_]', '_', name)
        clean = re.sub(r'_+', '_', clean).strip('_').lower()
        if not clean or clean[0].isdigit():
            clean = 'tbl_' + clean
        return clean

    # ── Clipboard/Save helpers ────────────────────────────────────────────────

    def _copy_ddl(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.ddl_text.get("1.0", "end-1c"))
        self.status_var.set("DDL copied to clipboard")

    def _copy_dml(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.dml_text.get("1.0", "end-1c"))
        self.status_var.set("DML copied to clipboard")

    def _save_text(self, text_widget, default_name: str, file_type_name: str, file_type_ext: str):
        file_path = filedialog.asksaveasfilename(
            defaultextension=Path(default_name).suffix,
            initialfile=default_name,
            filetypes=[(file_type_name, file_type_ext)]
        )
        if file_path:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(text_widget.get("1.0", "end-1c"))
            self.status_var.set(f"Saved: {Path(file_path).name}")

    def _save_ddl(self):
        self._save_text(self.ddl_text, f"{self.table_name_var.get()}_ddl.sql", "SQL files", "*.sql")

    def _save_dml(self):
        self._save_text(self.dml_text, f"{self.table_name_var.get()}_dml.sql", "SQL files", "*.sql")

    def _save_csv(self):
        import csv
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=f"{self.table_name_var.get()}_staged.csv",
            filetypes=[("CSV files", "*.csv")]
        )
        if file_path:
            headers, rows = self.cleaner.get_staged_csv_rows()
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
            self.status_var.set(f"Saved CSV: {Path(file_path).name}")

    def _save_dict_xlsx(self):
        if not self.schema_engine:
            Messagebox.show_warning("Generate DDL first.", "No Data")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"{self.table_name_var.get()}_dictionary.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not file_path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Dictionary"
            headers = ['Column Name', 'Original Name', 'Data Type', 'Nullable',
                       'Null %', 'Is PK', 'Is FK Hint', 'Sample Values', 'Description']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for col in self.schema_engine.generate_data_dictionary():
                ws.append([col['column_name'], col['original_name'], col['data_type'],
                            "Yes" if col['nullable'] else "No", col['null_percentage'],
                            "Yes" if col['is_primary_key'] else "No",
                            "Yes" if col['is_foreign_key_hint'] else "No",
                            col['sample_values'], col['description']])
            for col in ws.columns:
                max_length = max((len(str(cell.value or '')) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            wb.save(file_path)
            self.status_var.set(f"Saved dictionary: {Path(file_path).name}")
        except ImportError:
            Messagebox.show_warning("openpyxl required. Install: pip install openpyxl", "Missing Dependency")

    def _save_dict_csv(self):
        if not self.schema_engine:
            Messagebox.show_warning("Generate DDL first.", "No Data")
            return
        import csv
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=f"{self.table_name_var.get()}_dictionary.csv",
            filetypes=[("CSV files", "*.csv")]
        )
        if file_path:
            dictionary = self.schema_engine.generate_data_dictionary()
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=dictionary[0].keys())
                writer.writeheader()
                writer.writerows(dictionary)
            self.status_var.set(f"Saved dictionary CSV: {Path(file_path).name}")

    def _export_styled_excel(self):
        if not self.schema_engine or not self.current_file:
            Messagebox.show_warning("Generate DDL first.", "No Data")
            return
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"{self.table_name_var.get()}_complete.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not file_path:
                return
            self.status_var.set("Creating styled workbook...")
            self.root.update_idletasks()
            saved_path = self.formatted_exporter.export_full_workbook(
                file_path=file_path,
                table_name=self.table_name_var.get(),
                headers=self.headers,
                cleaned_rows=self.cleaned_rows,
                data_dictionary=self.schema_engine.generate_data_dictionary(),
                cleaning_audit=self.cleaner.get_audit_log(),
                schema_columns=self.schema_engine.columns,
                column_map=self.schema_engine.get_column_name_map()
            )
            self.status_var.set(f"Exported: {Path(saved_path).name}")
            Messagebox.show_info(
                f"Saved to:\n{saved_path}\n\n4 sheets: Cleaned Data, Dictionary, Audit, Mapping",
                "Export Complete"
            )
        except Exception as e:
            self.status_var.set(f"Export error: {str(e)}")
            Messagebox.show_error(f"Export failed:\n{str(e)}", "Error")

    # ── Power BI ──────────────────────────────────────────────────────────────

    def _powerbi_export(self):
        if not self.schema_engine or not self.current_file:
            Messagebox.show_warning("Generate DDL first.", "No Data")
            return
        try:
            self.status_var.set("Exporting Power BI artifacts...")
            self.root.update_idletasks()
            exporter = PowerBIExporter()
            artifacts = exporter.export_all(
                table_name=self.table_name_var.get(),
                source_file=self.current_file,
                headers=self.headers,
                cleaned_rows=self.cleaned_rows,
                column_map=self.schema_engine.get_column_name_map(),
                cleaning_audit=self.cleaner.get_audit_log(),
                schema_columns=self.schema_engine.columns,
                dialect=self.dialect_var.get()
            )
            msg = ("Power BI Export Complete!\n\n"
                   f"Staged CSV:\n  {artifacts.get('staged_csv', 'N/A')}\n\n"
                   f"M Code (.pq):\n  {artifacts.get('m_code', 'N/A')}\n\n"
                   f"MERGE SQL:\n  {artifacts.get('merge_sql', 'N/A')}\n\n"
                   f"README:\n  {artifacts.get('readme', 'N/A')}\n\n"
                   f"Audit Log:\n  {artifacts.get('audit_log', 'N/A')}")
            Messagebox.show_info(msg, "Export Complete")
            self.status_var.set(f"Power BI export complete — {len(artifacts)} artifacts")
        except Exception as e:
            self.status_var.set(f"Export error: {str(e)}")
            Messagebox.show_error(f"Export failed:\n{str(e)}", "Error")

    def _preview_m_code(self):
        if not self.schema_engine or not self.current_file:
            Messagebox.show_warning("Generate DDL first.", "No Data")
            return
        try:
            from bi.m_code_generator import MCodeGenerator
            m_gen = MCodeGenerator(self.table_name_var.get(), self.dialect_var.get())
            m_code = m_gen.generate(
                source_file=self.current_file,
                headers=self.headers,
                column_map=self.schema_engine.get_column_name_map(),
                cleaning_audit=self.cleaner.get_audit_log(),
                schema_columns=self.schema_engine.columns
            )
            popup = ttk.Toplevel(self.root)
            popup.title("Power Query M Code Preview")
            popup.geometry("800x600")
            text = ScrolledText(popup, wrap="word", font=("Courier New", 11))
            text.pack(fill=BOTH, expand=True, padx=10, pady=10)
            text.insert("1.0", m_code)
            text.config(state="disabled")
            ttk.Button(
                popup, text="Copy to Clipboard",
                command=lambda: [popup.clipboard_clear(), popup.clipboard_append(m_code)]
            ).pack(pady=10)
            self.status_var.set("M Code preview opened")
        except Exception as e:
            self.status_var.set(f"Preview error: {str(e)}")

    # ── Dashboard ─────────────────────────────────────────────────────────────

    def _launch_dashboard(self):
        if not self.duckdb or not self.schema_engine:
            Messagebox.show_warning("Generate DDL first.", "No Data")
            return
        try:
            self.status_var.set("Starting dashboard server...")
            self.root.update_idletasks()

            suggester = ChartSuggester(self.schema_engine.columns)
            server = DashboardServer(
                duckdb_manager=self.duckdb,
                chart_suggester=suggester
            )
            url = server.start(open_browser=True)
            self.status_var.set(f"Dashboard running at {url}")
            Messagebox.show_info(
                f"Dashboard opened in your browser.\n\n{url}\n\nClose DataForge to stop the server.",
                "Dashboard Ready"
            )
        except Exception as e:
            self.status_var.set(f"Dashboard error: {str(e)}")
            Messagebox.show_error(f"Dashboard failed:\n{str(e)}", "Error")

    # ── Projects ──────────────────────────────────────────────────────────────

    def _save_project(self):
        if not self.schema_engine:
            Messagebox.show_warning("Generate DDL first before saving.", "No Data")
            return
        data = {
            'table_name': self.table_name_var.get(),
            'dialect': self.dialect_var.get(),
            'nullable': self.nullable_var.get(),
            'source_file': self.current_file or '',
            'source_sheet': self.sheet_var.get(),
            'row_count': len(self.cleaned_rows),
            'column_count': len(self.headers),
            'schema': {'columns': self.schema_engine.columns},
            'cleaning_recipe': {},
            'audit_log': self.cleaner.get_audit_log(),
            'ddl_output': self.ddl_text.get("1.0", "end-1c"),
            'dml_output': self.dml_text.get("1.0", "end-1c"),
            'data_dictionary': self.schema_engine.generate_data_dictionary(),
        }
        project_id = self.project_manager.save_project(
            f"{self.table_name_var.get()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            data
        )
        self.status_var.set(f"Project saved (ID: {project_id})")

    def _show_recent(self):
        projects = self.project_manager.list_projects(10)
        if not projects:
            Messagebox.show_info("No saved projects yet.", "Recent Projects")
            return
        msg = "Recent Projects:\n\n"
        for p in projects:
            msg += f"  {p['id']}: {p['name']} ({p['dialect']}, {p['row_count']} rows)\n"
        Messagebox.show_info(msg, "Recent Projects")

    def _load_preferences(self):
        dialect = self.project_manager.get_preference('default_dialect', 'snowflake')
        self.dialect_var.set(dialect)

    def run(self):
        self.root.mainloop()


def main():
    app = DataForgeApp()
    app.run()


if __name__ == "__main__":
    main()
