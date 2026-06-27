"""
DataForge Formatted Excel Exporter

Exports a styled multi-sheet Excel workbook with:
- Cleaned Data
- Data Dictionary
- Cleaning Audit
- Column Mapping
"""

from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime


class FormattedExcelExporter:
    """Exports styled Excel workbooks from DataForge project data."""

    def export_full_workbook(self, file_path: str, table_name: str,
                              headers: List[str], cleaned_rows: List[List[Any]],
                              data_dictionary: List[Dict[str, Any]],
                              cleaning_audit: List[Dict[str, Any]],
                              schema_columns: List[Dict[str, Any]],
                              column_map: Dict[str, str]) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl required. Install: pip install openpyxl")

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        alt_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header_row(ws, col_count):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            ws.row_dimensions[1].height = 20

        def auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or '')))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # Sheet 1: Cleaned Data
        ws1 = wb.active
        ws1.title = "Cleaned Data"
        clean_headers = [c['clean_name'] for c in schema_columns] if schema_columns else headers
        ws1.append(clean_headers)
        style_header_row(ws1, len(clean_headers))

        for i, row in enumerate(cleaned_rows):
            str_row = ['' if v is None else str(v) for v in row]
            ws1.append(str_row)
            if i % 2 == 1:
                for cell in ws1[i + 2]:
                    cell.fill = alt_fill

        ws1.freeze_panes = "A2"
        auto_width(ws1)

        # Sheet 2: Data Dictionary
        ws2 = wb.create_sheet("Data Dictionary")
        dict_headers = ['Column Name', 'Original Name', 'Data Type', 'Nullable',
                        'Null %', 'Is PK', 'Is FK Hint', 'Sample Values', 'Description']
        ws2.append(dict_headers)
        style_header_row(ws2, len(dict_headers))

        for col in data_dictionary:
            ws2.append([
                col['column_name'],
                col['original_name'],
                col['data_type'],
                "Yes" if col['nullable'] else "No",
                col['null_percentage'],
                "Yes" if col['is_primary_key'] else "No",
                "Yes" if col['is_foreign_key_hint'] else "No",
                col['sample_values'],
                col['description'],
            ])

        auto_width(ws2)

        # Sheet 3: Cleaning Audit
        ws3 = wb.create_sheet("Cleaning Audit")
        audit_headers = ['Timestamp', 'Action', 'Description', 'Rows Affected']
        ws3.append(audit_headers)
        style_header_row(ws3, len(audit_headers))

        for entry in cleaning_audit:
            ws3.append([
                entry.get('timestamp', ''),
                entry.get('action', ''),
                entry.get('description', ''),
                entry.get('rows_affected', 0),
            ])

        auto_width(ws3)

        # Sheet 4: Column Mapping
        ws4 = wb.create_sheet("Column Mapping")
        ws4.append(['Original Name', 'Clean Name'])
        style_header_row(ws4, 2)

        for orig, clean in column_map.items():
            ws4.append([orig, clean])

        auto_width(ws4)

        wb.save(file_path)
        return file_path
