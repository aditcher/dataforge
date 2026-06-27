"""
DataForge File Handler

Reads Excel (.xlsx, .xls) and CSV files into normalized
header + row format for the DataForge pipeline.
"""

import os
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional


class FileHandler:
    """Handles reading Excel and CSV files."""

    def read_file(self, file_path: str) -> Dict[str, Any]:
        """
        Read file metadata without loading all data.
        Returns file info dict.
        """
        path = Path(file_path)
        ext = path.suffix.lower()
        file_size = path.stat().st_size / (1024 * 1024)  # MB

        if ext in ('.xlsx', '.xls'):
            return self._read_excel_meta(file_path, file_size)
        elif ext == '.csv':
            return self._read_csv_meta(file_path, file_size)
        else:
            raise ValueError(f"Unsupported file type: {ext}. Use .xlsx, .xls, or .csv")

    def _read_excel_meta(self, file_path: str, file_size: float) -> Dict[str, Any]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            default_sheet = sheets[0] if sheets else 'Sheet1'

            ws = wb[default_sheet]
            headers = []
            row_count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else f'col_{j}' for j, c in enumerate(row)]
                else:
                    if any(c is not None for c in row):
                        row_count += 1
            wb.close()

            return {
                'file_type': 'excel',
                'file_size_mb': round(file_size, 2),
                'sheets': sheets,
                'default_sheet': default_sheet,
                'total_columns': len(headers),
                'total_rows': row_count,
                'headers': headers,
            }
        except ImportError:
            raise ImportError("openpyxl is required to read Excel files. Install: pip install openpyxl")

    def _read_csv_meta(self, file_path: str, file_size: float) -> Dict[str, Any]:
        import csv
        import chardet

        with open(file_path, 'rb') as f:
            raw = f.read(10000)
            detected = chardet.detect(raw)
            encoding = detected.get('encoding', 'utf-8') or 'utf-8'

        with open(file_path, 'r', encoding=encoding, errors='replace') as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            row_count = sum(1 for _ in reader)

        return {
            'file_type': 'csv',
            'file_size_mb': round(file_size, 2),
            'sheets': ['Sheet1'],
            'default_sheet': 'Sheet1',
            'total_columns': len(headers),
            'total_rows': row_count,
            'headers': headers,
            'encoding': encoding,
        }

    def load_sheet(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[List[str], List[List[Any]]]:
        """
        Load full data from a file sheet.
        Returns (headers, rows).
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext in ('.xlsx', '.xls'):
            return self._load_excel_sheet(file_path, sheet_name)
        elif ext == '.csv':
            return self._load_csv(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    def _load_excel_sheet(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[List[str], List[List[Any]]]:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        headers = []
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else f'col_{j}' for j, c in enumerate(row)]
            else:
                if any(c is not None for c in row):
                    rows.append(list(row))
        wb.close()
        return headers, rows

    def _load_csv(self, file_path: str) -> Tuple[List[str], List[List[Any]]]:
        import csv
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw = f.read(10000)
            detected = chardet.detect(raw)
            encoding = detected.get('encoding', 'utf-8') or 'utf-8'
        except ImportError:
            encoding = 'utf-8'

        headers = []
        rows = []
        with open(file_path, 'r', encoding=encoding, errors='replace') as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            rows = [list(row) for row in reader if any(c.strip() for c in row)]
        return headers, rows

    def get_all_sheets(self, file_path: str) -> Dict[str, Tuple[List[str], List[List[Any]]]]:
        """Load all sheets from an Excel file."""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else f'col_{j}' for j, c in enumerate(row)]
                else:
                    if any(c is not None for c in row):
                        rows.append(list(row))
            result[sheet_name] = (headers, rows)
        wb.close()
        return result
